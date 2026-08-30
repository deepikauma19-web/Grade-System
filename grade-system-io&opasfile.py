"""
Reads student names and subject marks from an Excel file, then prints
each student's total, average, and letter grade, and saves the results
to a new Excel file (grade_report.xlsx).

Expected sheet layout (first row = headers, first column = name):

    Name     | Maths | Science | English | History
    Alice    | 95    | 88      | 91      | 85
    Bob      | 72    | 68      | 75      | 70

Requires: openpyxl  (pip install openpyxl --break-system-packages)
"""

from openpyxl import load_workbook, Workbook


def get_grade(average):
    if average >= 90:
        return "A"
    elif average >= 80:
        return "B"
    elif average >= 70:
        return "C"
    elif average >= 60:
        return "D"
    else:
        return "E"


def main():
    filename = "sample_marks.xlsx"   # change this to your file name

    wb = load_workbook(filename)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    subjects = rows[0][1:]   # header row, skipping the "Name" column

    print("STUDENT GRADE REPORT")
    print("=" * 40)

    # New workbook to save the results into
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.append(["Name"] + list(subjects) + ["Total", "Average", "Grade"])

    for row in rows[1:]:
        name = row[0]
        marks = row[1:]

        total = sum(marks)
        average = total / len(marks)
        grade = get_grade(average)

        print(f"\n{name}")
        for subject, mark in zip(subjects, marks):
            print(f"  {subject}: {mark}")
        print(f"  Total: {total}")
        print(f"  Average: {average:.2f}")
        print(f"  Grade: {grade}")

        out_ws.append([name] + list(marks) + [total, round(average, 2), grade])

    out_wb.save("grade_report.xlsx")
    print("\nSaved results to grade_report.xlsx")


if __name__ == "__main__":
    main()